from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
import json
import openpyxl
from django.utils import timezone


# ─── Conversión de fecha serial de Excel ───────────────────────────────────────
def excel_serial_to_datetime(serial):
    """
    Convierte un número serial de Excel (float) a datetime con zona horaria.
    También acepta objetos datetime nativos o strings ISO 8601.
    Retorna None si el valor no puede convertirse.
    """
    if serial is None:
        return None
    try:
        if isinstance(serial, datetime):
            return timezone.make_aware(serial) if timezone.is_naive(serial) else serial

        serial = float(serial)
        base = datetime(1899, 12, 30)
        dt = base + timedelta(days=serial)
        return timezone.make_aware(dt)
    except (ValueError, TypeError):
        try:
            dt = datetime.fromisoformat(str(serial))
            return timezone.make_aware(dt) if timezone.is_naive(dt) else dt
        except Exception:
            return None


# ─── Normalización del payload JSON ────────────────────────────────────────────
def normalize_payload(raw):
    """
    Normaliza el payload JSON de route_payload.
    Soporta variaciones de nombres de campos:
      - 'direccion' o 'address'
      - 'latitud' o 'lat'
      - 'longitud' o 'lon'
    Convierte coordenadas a float (string o número).
    """
    if not raw:
        return None

    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except json.JSONDecodeError:
            return None

    if not isinstance(raw, dict):
        return None

    latitud = raw.get('latitud') or raw.get('lat')
    longitud = raw.get('longitud') or raw.get('lon')
    direccion = raw.get('direccion') or raw.get('address')
    piezas = raw.get('piezas', [])
    primer_peso = piezas[0].get('peso') if piezas else None

    def safe_float(val):
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    return {
        'id_punto':    raw.get('idPunto'),
        'direccion':   direccion,
        'latitud':     safe_float(latitud),
        'longitud':    safe_float(longitud),
        'primer_peso': safe_float(primer_peso),
    }


# ─── Validación de coordenadas (rango Colombia) ────────────────────────────────
def validate_coordinates(lat, lon):
    """
    Valida que las coordenadas estén dentro del rango geográfico de Colombia:
      Latitud:  -4.3 a 13.5
      Longitud: -82.0 a -56.0
    Acepta valores numéricos o strings numéricos.
    Retorna (True, None) si son válidas, (False, motivo) si no lo son.
    """
    if lat is None or lon is None:
        return False, 'Coordenadas fuera del rango geográfico válido para Colombia'
    try:
        lat = float(lat)
        lon = float(lon)
    except (ValueError, TypeError):
        return False, 'Coordenadas no son numéricas'

    if not (-4.3 <= lat <= 13.5):
        return False, (
            f'Coordenadas fuera del rango geográfico válido para Colombia '
            f'(latitud {lat} fuera de [-4.3, 13.5])'
        )
    if not (-82.0 <= lon <= -56.0):
        return False, (
            f'Coordenadas fuera del rango geográfico válido para Colombia '
            f'(longitud {lon} fuera de [-82.0, -56.0])'
        )

    return True, None


# ─── Mapeo flexible de encabezados ──────────────────────────────────────────────
def normalize_header(header):
    if not header:
        return ""
    return str(header).lower().replace("_", "").replace(" ", "").strip()


def get_header_map(headers):
    """Crea un mapa de encabezados encontrados vs nombres canónicos esperados."""
    mapping = {
        'idroute':           ['idroute', 'idruta', 'id', 'id_route', 'id_ruta'],
        'idoficinid':        ['idoficinaorigen', 'idoficina', 'oficina', 'id_oficina', 'oficina_id'],
        'origin':            ['origin', 'origen', 'punto_origen', 'salida'],
        'destination':       ['destination', 'destino', 'punto_destino', 'llegada'],
        'distance_km':       ['distancekm', 'distanciakm', 'distancia', 'km'],
        'priority':          ['priority', 'prioridad', 'nivel_prioridad'],
        'time_window_start': ['timewindowstart', 'ventanainicio', 'inicio', 'desde'],
        'time_window_end':   ['timewindowend', 'ventanafin', 'fin', 'hasta'],
        'status':            ['status', 'estado', 'fase'],
        'created_at':        ['createdat', 'fecharegistro', 'fecha', 'creado', 'timestamp'],
    }

    found_map = {}
    normalized_headers = [normalize_header(h) for h in headers]

    for target, aliases in mapping.items():
        for i, h in enumerate(normalized_headers):
            if h in aliases:
                found_map[target] = headers[i]
                break
    return found_map


# ─── Parser principal del Excel ────────────────────────────────────────────────
def parse_excel(file):
    """
    Parsea un archivo Excel (.xlsx) con las 6 hojas del dataset:
      - oficina_org, priorities_ref, poblacion_cor: tablas maestras (opcionales)
      - routes: tabla principal de rutas
      - route_payload: payloads JSON por ruta (con campos inconsistentes entre filas)
      - execution_logs: se ignora (se reimportan desde el sistema)

    Retorna:
      {
        'valid_rows': [...],    # filas válidas listas para insertar
        'errors': [...],        # {row, field, value, reason}
        'has_master_data': {...} # datos de tablas maestras
      }
    """
    valid_rows = []
    errors = []

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return {
            'valid_rows': [],
            'errors': [{'row': 0, 'field': 'file', 'value': None, 'reason': f'No se pudo leer el archivo: {e}'}],
            'has_master_data': {},
        }

    # Determinar hoja de rutas
    if 'routes' in wb.sheetnames:
        sheet_name = 'routes'
    else:
        sheet_name = wb.sheetnames[0]

    # ── Cargar hojas maestras ──────────────────────────────────────────────────
    oficinas_data = []
    priorities_data = []
    poblaciones_data = []

    for sname in wb.sheetnames:
        norm_name = normalize_header(sname)
        ws = wb[sname]

        if norm_name in ['oficinaorg', 'oficina', 'oficinas']:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    oficinas_data.append({'id': str(row[0]).strip(), 'nombre': row[1]})

        elif norm_name in ['prioritiesref', 'priority', 'prioridades']:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    priorities_data.append({'id': str(row[0]).strip(), 'nombre': row[1]})

        elif norm_name in ['poblacioncor', 'poblacion', 'puntos']:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    poblaciones_data.append({
                        'id':     str(row[0]).strip(),
                        'ciudad': row[1],
                        'lat':    row[2],
                        'lon':    row[3],
                    })

    # ── Cargar payloads de route_payload ──────────────────────────────────────
    payloads = {}
    payload_sheet = next(
        (s for s in wb.sheetnames if normalize_header(s) in ['routepayload', 'payload', 'datos']),
        None
    )
    if payload_sheet:
        ws_payload = wb[payload_sheet]
        for row in ws_payload.iter_rows(min_row=2, values_only=True):
            if len(row) >= 2:
                id_route_raw, raw_payload = row[0], row[1]
                if id_route_raw is not None:
                    key = str(id_route_raw).strip()
                    payloads[key] = normalize_payload(raw_payload)

    # ── Procesar hoja de rutas ─────────────────────────────────────────────────
    ws_routes = wb[sheet_name]

    header_row = next(ws_routes.iter_rows(min_row=1, max_row=1), None)
    if not header_row:
        return {
            'valid_rows': [],
            'errors': [{'row': 0, 'field': 'sheet', 'value': None, 'reason': 'La hoja de rutas está vacía'}],
            'has_master_data': {},
        }

    headers = [cell.value for cell in header_row]
    h_map = get_header_map(headers)

    VALID_STATUSES = {'READY', 'PENDING', 'EXECUTED', 'FAILED'}

    for row_idx, row in enumerate(ws_routes.iter_rows(min_row=2, values_only=True), start=2):
        # Omitir filas completamente vacías
        if all(v is None for v in row):
            continue

        row_data = dict(zip(headers, row))
        row_errors = []

        # ── Extraer campos usando mapa flexible ───────────────────────────────
        id_route_raw       = row_data.get(h_map.get('idroute'))
        id_oficina         = row_data.get(h_map.get('idoficinid'))
        origin             = row_data.get(h_map.get('origin'))
        destination        = row_data.get(h_map.get('destination'))
        distance_km_raw    = row_data.get(h_map.get('distance_km'))
        priority_raw       = row_data.get(h_map.get('priority'))
        time_window_start  = row_data.get(h_map.get('time_window_start'))
        time_window_end    = row_data.get(h_map.get('time_window_end'))
        status_raw         = row_data.get(h_map.get('status'))
        created_at         = row_data.get(h_map.get('created_at'))

        # ── Validar origin ────────────────────────────────────────────────────
        if not origin or not str(origin).strip():
            row_errors.append({
                'row': row_idx, 'field': 'origin',
                'value': origin, 'reason': 'El origen no puede estar vacío',
            })
            origin = None
        else:
            origin = str(origin).strip()

        # ── Validar destination ───────────────────────────────────────────────
        if not destination or not str(destination).strip():
            row_errors.append({
                'row': row_idx, 'field': 'destination',
                'value': destination, 'reason': 'El destino no puede estar vacío',
            })
            destination = None
        else:
            destination = str(destination).strip()

        # ── Validar distance_km ───────────────────────────────────────────────
        distance_km = None
        if distance_km_raw is None or str(distance_km_raw).strip() == '':
            row_errors.append({
                'row': row_idx, 'field': 'distance_km',
                'value': distance_km_raw, 'reason': 'La distancia es obligatoria',
            })
        else:
            try:
                distance_km = float(distance_km_raw)
                if distance_km <= 0:
                    row_errors.append({
                        'row': row_idx, 'field': 'distance_km',
                        'value': distance_km_raw, 'reason': 'La distancia debe ser mayor que 0',
                    })
                    distance_km = None
            except (ValueError, TypeError):
                row_errors.append({
                    'row': row_idx, 'field': 'distance_km',
                    'value': distance_km_raw, 'reason': 'La distancia debe ser un número',
                })

        # ── Validar priority ──────────────────────────────────────────────────
        priority = None
        if priority_raw is None or str(priority_raw).strip() == '':
            row_errors.append({
                'row': row_idx, 'field': 'priority',
                'value': priority_raw, 'reason': 'La prioridad es obligatoria',
            })
        else:
            try:
                priority = int(priority_raw)
                if priority <= 0:
                    row_errors.append({
                        'row': row_idx, 'field': 'priority',
                        'value': priority_raw, 'reason': 'La prioridad debe ser un entero positivo',
                    })
                    priority = None
            except (ValueError, TypeError):
                row_errors.append({
                    'row': row_idx, 'field': 'priority',
                    'value': priority_raw, 'reason': 'La prioridad debe ser un entero',
                })

        # ── Validar status ────────────────────────────────────────────────────
        status = str(status_raw).upper().strip() if status_raw else 'READY'
        if status not in VALID_STATUSES:
            status = 'READY'

        # ── Convertir y validar fechas ────────────────────────────────────────
        tw_start = excel_serial_to_datetime(time_window_start)
        tw_end   = excel_serial_to_datetime(time_window_end)

        if tw_start is None:
            row_errors.append({
                'row': row_idx, 'field': 'time_window_start',
                'value': time_window_start, 'reason': 'Fecha/hora de inicio inválida o no convertible',
            })
        if tw_end is None:
            row_errors.append({
                'row': row_idx, 'field': 'time_window_end',
                'value': time_window_end, 'reason': 'Fecha/hora de fin inválida o no convertible',
            })
        if tw_start and tw_end and tw_start >= tw_end:
            row_errors.append({
                'row': row_idx, 'field': 'time_window_end',
                'value': time_window_end,
                'reason': 'time_window_start debe ser anterior a time_window_end',
            })
            tw_end = None  # invalidar

        # ── Validar coordenadas del payload ───────────────────────────────────
        id_route_key = str(id_route_raw).strip() if id_route_raw is not None else None
        payload = payloads.get(id_route_key) if id_route_key else None

        if payload:
            coord_ok, coord_err = validate_coordinates(payload.get('latitud'), payload.get('longitud'))
            if not coord_ok:
                row_errors.append({
                    'row': row_idx, 'field': 'coordenadas',
                    'value': f"{payload.get('latitud')}, {payload.get('longitud')}",
                    'reason': coord_err,
                })

        # ── Resultado de la fila ───────────────────────────────────────────────
        if row_errors:
            errors.extend(row_errors)
            continue

        # Generar id_route si falta (como string)
        if not id_route_key:
            id_route_key = f'AUTO-{row_idx}'

        # Normalizar id_oficina como string
        id_oficina_str = str(id_oficina).strip() if id_oficina is not None else None

        cr_at = excel_serial_to_datetime(created_at) or timezone.now()

        valid_rows.append({
            'id_route':          id_route_key,
            'id_oficina_id':     id_oficina_str,
            'origin':            origin,
            'destination':       destination,
            'distance_km':       Decimal(str(distance_km)),
            'priority':          priority,
            'time_window_start': tw_start,
            'time_window_end':   tw_end,
            'status':            status,
            'payload':           payload,
            'created_at':        cr_at,
        })

    return {
        'valid_rows': valid_rows,
        'errors': errors,
        'has_master_data': {
            'oficinas':    oficinas_data,
            'priorities':  priorities_data,
            'poblaciones': poblaciones_data,
        },
    }