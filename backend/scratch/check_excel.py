import openpyxl
import os

file_path = 'c:/Users/Gamer/logistics-routes-system/backend/data/dataset2.xlsx'

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
else:
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        print(f"Sheets: {wb.sheetnames}")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            print(f"Headers in {sheet_name}: {headers}")
    except Exception as e:
        print(f"Error reading file: {e}")
